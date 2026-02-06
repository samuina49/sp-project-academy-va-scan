"""
Professional Excel Report Generator for Vulnerability Scan Results
Enhanced with proper spacing, alignment, and official document formatting
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List
from io import BytesIO

from app.models.scan_models import FileScanResult


class ExcelReportGenerator:
    """Generate professional, well-formatted Excel reports for vulnerability scans"""
    
    # Enhanced color scheme for severity levels
    SEVERITY_COLORS = {
        'CRITICAL': 'DC143C',  # Crimson Red
        'HIGH': 'FF6347',      # Tomato
        'MEDIUM': 'FFA500',    # Orange
        'LOW': '32CD32',       # Lime Green
        'INFO': '4169E1'       # Royal Blue
    }
    
    # Border styles
    THICK_BORDER = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Vulnerability Report"
    
    def generate_report(
        self, 
        scan_results: List[FileScanResult],
        scan_metadata: dict = None
    ) -> BytesIO:
        """
        Generate professionally formatted Excel report
        
        Args:
            scan_results: List of file scan results
            scan_metadata: Optional metadata (scan_id, timestamp, etc.)
            
        Returns:
            BytesIO object containing Excel file
        """
        # Add title and metadata with spacing
        self._add_report_header(scan_metadata or {})
        
        # Add detailed findings table (skip summary)
        self._add_findings_table(scan_results)
        
        # Apply column widths and formatting
        self._apply_professional_formatting()
        
        # Save to BytesIO
        excel_file = BytesIO()
        self.wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file
    
    def _add_report_header(self, metadata: dict):
        """Add professionally formatted report header"""
        # Row 1: Main Title (Merged across all columns)
        self.ws.merge_cells('A1:H1')
        title_cell = self.ws['A1']
        title_cell.value = '🛡️ VULNERABILITY SCAN REPORT'
        title_cell.font = Font(size=20, bold=True, color='FFFFFF')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        title_cell.border = self.THICK_BORDER
        self.ws.row_dimensions[1].height = 45
        
        # Row 2: Blank spacer
        self.ws.row_dimensions[2].height = 20
        
        # Row 3-6: Metadata section (Labels in column A, Values merged B-D)
        metadata_row = 3
        
        # Scan ID
        label_cell = self.ws[f'A{metadata_row}']
        label_cell.value = 'Scan ID:'
        label_cell.font = Font(bold=True, size=12)
        label_cell.alignment = Alignment(horizontal='right', vertical='center')
        
        self.ws.merge_cells(f'B{metadata_row}:D{metadata_row}')
        value_cell = self.ws[f'B{metadata_row}']
        value_cell.value = metadata.get('scan_id', 'N/A')
        self._format_metadata_value_cell(f'B{metadata_row}')
        
        # Scan Date
        metadata_row += 1
        label_cell = self.ws[f'A{metadata_row}']
        label_cell.value = 'Scan Date:'
        label_cell.font = Font(bold=True, size=12)
        label_cell.alignment = Alignment(horizontal='right', vertical='center')
        
        self.ws.merge_cells(f'B{metadata_row}:D{metadata_row}')
        value_cell = self.ws[f'B{metadata_row}']
        value_cell.value = metadata.get('timestamp', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        self._format_metadata_value_cell(f'B{metadata_row}')
        
        # Total Files
        metadata_row += 1
        label_cell = self.ws[f'A{metadata_row}']
        label_cell.value = 'Total Files Scanned:'
        label_cell.font = Font(bold=True, size=12)
        label_cell.alignment = Alignment(horizontal='right', vertical='center')
        
        self.ws.merge_cells(f'B{metadata_row}:D{metadata_row}')
        value_cell = self.ws[f'B{metadata_row}']
        value_cell.value = metadata.get('total_files', 1)
        self._format_metadata_value_cell(f'B{metadata_row}')
        
        # Language
        metadata_row += 1
        label_cell = self.ws[f'A{metadata_row}']
        label_cell.value = 'Language:'
        label_cell.font = Font(bold=True, size=12)
        label_cell.alignment = Alignment(horizontal='right', vertical='center')
        
        self.ws.merge_cells(f'B{metadata_row}:D{metadata_row}')
        value_cell = self.ws[f'B{metadata_row}']
        value_cell.value = metadata.get('language', 'Multiple')
        self._format_metadata_value_cell(f'B{metadata_row}')
        
        # Set row heights for metadata
        for row in range(3, 7):
            self.ws.row_dimensions[row].height = 28
    
    def _format_metadata_value_cell(self, cell_ref):
        """Apply consistent formatting to metadata value cells"""
        cell = self.ws[cell_ref]
        cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
        cell.border = self.THIN_BORDER
        cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        cell.font = Font(size=11)
    
    def _add_findings_table(self, scan_results: List[FileScanResult]):
        """Add detailed findings table with professional formatting"""
        # Start right after metadata (row 9)
        start_row = 9
        self.ws.row_dimensions[start_row - 1].height = 20  # Spacer
        
        # Section title
        self.ws.merge_cells(f'A{start_row}:H{start_row}')
        title_cell = self.ws[f'A{start_row}']
        title_cell.value = '📋 VULNERABILITY FINDINGS'
        title_cell.font = Font(size=16, bold=True, color='FFFFFF')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color='C55A11', end_color='C55A11', fill_type='solid')
        title_cell.border = self.THICK_BORDER
        self.ws.row_dimensions[start_row].height = 35
        
        # Column headers
        row = start_row + 1
        headers = ['ID', 'File', 'Line', 'Severity', 'Vulnerability Type', 'CWE', 'OWASP', 'Description']
        for col_idx, header in enumerate(headers, 1):
            cell = self.ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(bold=True, color='FFFFFF', size=12)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.THICK_BORDER
        self.ws.row_dimensions[row].height = 40
        
        # Add findings data
        row += 1
        finding_num = 1
        for result in scan_results:
            for finding in result.findings:
                # Map severity
                severity_str = str(finding.severity).upper()
                if 'ERROR' in severity_str or 'CRITICAL' in severity_str:
                    severity = 'CRITICAL' if 'CRITICAL' in severity_str else 'HIGH'
                elif 'WARNING' in severity_str:
                    severity = 'MEDIUM'
                elif 'INFO' in severity_str or 'LOW' in severity_str:
                    severity = 'LOW'
                else:
                    severity = 'MEDIUM'
                
                # Clean vulnerability name
                vuln_name = finding.message.split('\n')[0] if finding.message else finding.rule_id
                if len(vuln_name) > 70:
                    vuln_name = vuln_name[:67] + '...'
                
                # Prepare data
                data = [
                    finding_num,
                    result.file_path,
                    finding.start_line,
                    severity,
                    vuln_name,
                    getattr(finding, 'cwe_id', None) or 'N/A',
                    getattr(finding, 'owasp_category', None) or 'N/A',
                    finding.message[:300] + '...' if finding.message and len(finding.message) > 300 else (finding.message or '')
                ]
                
                # Add data to cells
                for col_idx, value in enumerate(data, 1):
                    cell = self.ws.cell(row=row, column=col_idx, value=value)
                    cell.border = self.THIN_BORDER
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                    cell.font = Font(size=11)
                    
                    # Format specific columns
                    if col_idx == 1:  # ID column
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.font = Font(bold=True, size=11)
                        cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
                    elif col_idx == 3:  # Line column
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif col_idx == 4:  # Severity column
                        color = self.SEVERITY_COLORS.get(severity, 'FFFFFF')
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        cell.font = Font(color='FFFFFF', bold=True, size=11)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif col_idx in [6, 7]:  # CWE and OWASP
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Set row height based on content
                self.ws.row_dimensions[row].height = 60  # Tall enough for wrapped text
                row += 1
                finding_num += 1
    
    def _apply_professional_formatting(self):
        """Apply final professional formatting touches"""
        # Column widths (optimized for readability with better spacing)
        column_widths = {
            'A': 6,    # ID
            'B': 30,   # File (wider)
            'C': 8,    # Line
            'D': 14,   # Severity
            'E': 50,   # Vulnerability Type (wider)
            'F': 15,   # CWE
            'G': 35,   # OWASP (wider)
            'H': 80    # Description (much wider)
        }
        
        for col_letter, width in column_widths.items():
            self.ws.column_dimensions[col_letter].width = width
        
        # Freeze panes (freeze headers at row 11)
        self.ws.freeze_panes = 'A11'


def generate_excel_report(scan_results: List[FileScanResult], metadata: dict = None) -> BytesIO:
    """
    Convenience function to generate professional Excel report
    
    Args:
        scan_results: List of file scan results
        metadata: Optional scan metadata
        
    Returns:
        BytesIO containing Excel file
    """
    generator = ExcelReportGenerator()
    return generator.generate_report(scan_results, metadata)
