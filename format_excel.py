import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

input_file = r'c:\Users\user\Desktop\Project Final University Bon\sp-project-academy-va-scan\backend\data\training_dataset_cleaned.json'
output_file = r'c:\Users\user\Desktop\Project Final University Bon\sp-project-academy-va-scan\backend\data\training_dataset_formatted.xlsx'

with open(input_file, 'r', encoding='utf-8') as f:
    data = json.load(f)

# Parse data
rows = []
for item in data:
    metadata = item.get('metadata', {})
    
    vuln_types = metadata.get('vulnerability_types', [])
    if isinstance(vuln_types, list):
        if len(vuln_types) > 0 and isinstance(vuln_types[0], dict):
            vuln_str = ', '.join([v.get('type', '') for v in vuln_types])
        else:
            vuln_str = ', '.join(vuln_types)
    else:
        vuln_str = str(vuln_types)
        
    if not vuln_str:
        vuln_str = 'Safe (No Vulnerability)'
        
    rows.append({
        'Status': 'Vulnerable' if item.get('label', 0) == 1 else 'Safe',
        'Vulnerability Type': vuln_str.title().replace('_', ' '),
        'Source Code': item.get('code', '').strip(),
        'File Source': item.get('file_path', ''),
        'Dataset Origin': metadata.get('source', '').title().replace('_', ' ')
    })

df = pd.DataFrame(rows)

# Create workbook
wb = Workbook()
wb.remove(wb.active) # remove default sheet

# Common styles
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def style_sheet(ws):
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        
    # Auto adjust columns
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 80
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 20
    
    # Wrap text for code
    for row in ws.iter_rows(min_row=2, max_col=5):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = border

# 1. Summary Sheet
summary = df['Vulnerability Type'].value_counts().reset_index()
summary.columns = ['Vulnerability Type', 'Count']
ws_sum = wb.create_sheet(title="Dataset Summary")
for r in dataframe_to_rows(summary, index=False, header=True):
    ws_sum.append(r)

for cell in ws_sum[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

ws_sum.column_dimensions['A'].width = 35
ws_sum.column_dimensions['B'].width = 15

# 2. Main Dataset (First 1000 rows as 'Preview' to not overload Excel if needed, but we can put all)
# Since 28k rows is small enough for Excel, we'll put it all in one "Full Dataset" sheet.
ws_full = wb.create_sheet(title="Full Dataset")
for r in dataframe_to_rows(df, index=False, header=True):
    ws_full.append(r)
style_sheet(ws_full)

# 3. Create sample sheets for each top vulnerability to make it easy to copy-paste to Word
top_vulns = [v for v in df['Vulnerability Type'].unique() if v != 'Safe (No Vulnerability)'][:5]
for vuln in top_vulns:
    df_sub = df[df['Vulnerability Type'] == vuln].head(50) # Take 50 samples per vuln for the report
    
    sheet_name = str(vuln)[:31] # Excel sheet name limit
    ws_sub = wb.create_sheet(title=f"Sample - {sheet_name}")
    for r in dataframe_to_rows(df_sub, index=False, header=True):
        ws_sub.append(r)
    style_sheet(ws_sub)

# Save
wb.save(output_file)
print(f"Excel formatted saved successfully to {output_file}")
